from distutils.log import ERROR, info
import logging
from openpyxl import load_workbook
import sqlite3 
from telegram.error import  BadRequest
from telegram import (
                      Bot,
                      Update, 
                      ReplyKeyboardMarkup,
                      KeyboardButton)

from telegram.ext import  (BaseFilter,
                           Updater,
                           CallbackContext,
                           Dispatcher,
                           CommandHandler,
                           MessageHandler,
                           Filters,
                           ConversationHandler,
                           MessageFilter
                           )

connection = sqlite3.connect("base.sqlite3",check_same_thread=False)
query = """INSERT INTO users(name, surname, father_name, user_id, phone_number, region, district, school, class) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)"""
cursor = connection.cursor()

wb = load_workbook('users.xlsx')

work_sheet = wb.active # Get active sheet
# work_sheet['A1'] = "Familiya"
# work_sheet['B1'] = 'Ism'
# work_sheet['C1'] = 'Sharifi'
# work_sheet['E1'] = 'Telefon raqami'
# work_sheet['D1'] = 'User_id'
# work_sheet['F1'] = 'Viloyat'
# work_sheet['G1'] = 'Tuman'
# work_sheet['H1'] = 'Maktab raqami'
# work_sheet['I1'] = 'Sinf raqami'
#  # owner_id
# wb.save('users.xlsx')
owner_id = "640077553"
    
updater = Updater(token="5154057051:AAFQ3gJTTBY0YwX5sqq65yzSoNH-QWFJdJ4")
logging.basicConfig(level=ERROR)

class FilterFullname(MessageFilter):
    
    def filter(self, message):
        
        fullname = message.text
        return (len(fullname.split()) > 1) and  fullname.split()[0][0].isupper() and fullname.split()[1][0].isupper()
filter_fullname =  FilterFullname()

class FilterNumber(MessageFilter):
    
    def filter(self, message):
        phone_number = message.text
        return (len(phone_number) == 14 or len(phone_number) == 9) or phone_number[1:].isnumeric()
filter_number = FilterNumber()
 

REGIONS = ReplyKeyboardMarkup([
    ["Toshkent",  "Andijon"],
    ["Farg'ona", "Namangan", "Navoiy"],
    ["Toshkent shahri", "Buxoro", "Qashqadaryo"],
    ["Samarqand", "Sirdaryo", "Qoraqalpog'iston"],
    ["Surxondaryo", "Jizzax", "Xorazm"],
    ['⬅️ Orqaga']
])
BUTTON2 = ReplyKeyboardMarkup([ 
                ['1','2','3','4'],
                ['5','6','7','8'],
                ['9','10','11']

], resize_keyboard=True)

regions = ["Toshkent",  "Andijon", "Farg'ona", "Namangan", "Navoiy",
           "Nukus", "Buxoro", "Qashqadaryo","Samarqand", "Sirdaryo", "Zarafshon",
           "Surxondaryo", "Jizzax", "Xorazm"]


cursor.execute('''SELECT * FROM districts''')
districts = [ls[1] for ls in cursor]


BUTTON = ReplyKeyboardMarkup([["📝 Ro'yxatdan o'tish"]], resize_keyboard=True)

def start_bot(update: Update, context: CallbackContext):
    update.effective_chat.send_action('typing') 
    context.user_data['id'] = update.effective_user.id                #user_id
    cursor.execute('''SELECT user_id FROM users''')
    id_list = [int(id[0]) for id in cursor]
    print("IDLIST", id_list)
    if update.effective_user.id not in id_list:
        # return 'main_menu'
        update.message.reply_text("Xush kelibsiz! Ushbu bot orqali Yoshlar ishlari agentligi, Yoshlar ittifoqi va “Mening yurtim” telekanali hamkorligida o'tkazilayotgan onlayn olimpiadada ishtirok etishingiz mumkin!")
        update.message.reply_html("Olimpiyadada ishtirok etish uchun <b> Ro'yxatdan o'tish </b> tugmasini bosing.", reply_markup=BUTTON)
        return 'fullname'
    else:
        update.message.reply_html("‹‹ ʙɪʟɪᴍᴛᴇsᴛʙᴏᴛ ››", reply_markup=ReplyKeyboardMarkup([["🔠 Test Ishlash"], [" ℹ️ Ma'lumot", "👤 Profilim"]], resize_keyboard=True))

# def start(update:Update, context: CallbackContext):
#     update.message.reply_html("‹‹ ʙɪʟɪᴍᴛᴇsᴛʙᴏᴛ ››", reply_markup=ReplyKeyboardMarkup([["🔠 Test Ishlash"], [" ℹ️ Ma'lumot", "👤 Profilim"]], resize_keyboard=True))


 


     
def get_fullname1(update: Update, context: CallbackContext):
    update.effective_chat.send_action('typing')
    
    update.message.reply_html("<b>F.I.SH.</b>ni kiriting")
    return 'phone_number'
 
def get_fullname2(update: Update, context: CallbackContext):
    update.effective_chat.send_action('typing')

    update.message.reply_text("Familiya, ism, sharifni to'g'ri kiriting\n\nMasalan, Aliyev Vali ")
    return 'phone_number'

def get_contact(update: Update, context: CallbackContext):
    
    data = update.message.text
    if (len(data.split()) > 1) and  data.split()[0][0].isupper() and data.split()[1][0].isupper() :
        
        context.user_data['surname'] = (update.message.text).split()[0]    #user_surname
        print(context.user_data['surname'])
        context.user_data['name'] = (update.message.text).split()[1]         #user_name
        if len((update.message.text).split()) == 3:
            context.user_data['father_name'] = (update.message.text).split()[2]      #user_father_name
        else:
            context.user_data['father_name'] = ""
        key = ReplyKeyboardMarkup([[KeyboardButton("📱 Telefon raqamni jo'natish", request_contact=True)]],
                                 resize_keyboard=True)
        update.message.reply_text("Tugmani bosib telefon raqamingizni yuboring \nYoki raqamingizni kiriting.", reply_markup=key)
        return 'region'
    
    return 'repeat_name'
 
def get_region(update: Update, context: CallbackContext):
    data = update.message.text
    print("DATA", data)
    if data == None:
        context.user_data['number'] = data  #phone_number
        update.message.reply_text("Viloyatni tanlang", reply_markup=REGIONS)
        return 'district'
    elif (len(data) == 14 or len(data) == 9) or data[1:].isnumeric():
        context.user_data['number'] = data  #phone_number
        update.message.reply_text("Viloyatni tanlang", reply_markup=REGIONS)
        return 'district'
    return 'phone_number'

def get_district(update: Update, context: CallbackContext):
    region = update.message.text
    context.user_data['region'] = region
    cursor.execute(f'''SELECT regionid FROM regions WHERE name = "{region}" ''')
    for cur in cursor:
        id = int(cur[0])
    # print("PRINT:", id)
    cursor.execute(f'''SELECT name FROM districts WHERE district_regionid = "{id}" ''')
   
    update.message.reply_text(region)
    BUTTON3 = ReplyKeyboardMarkup([list([cur[0]]) for cur in cursor]+[['⬅️ Orqaga']])
    update.message.reply_text("Tumanni tanlang", reply_markup=BUTTON3)
    return 'school_number'

def get_school_number(update: Update, context:CallbackContext):
    context.user_data['district'] = update.message.text
    update.message.reply_text("Maktabingiz raqamini kiriting")
    return 'class_number'

def get_class_number(update: Update, context: CallbackContext):
    data = update.message.text
    if data.isnumeric():
        context.user_data['school_num'] = update.message.text
        update.message.reply_text("Sinfingiz raqamini kiriting", reply_markup=BUTTON2)
        return 'main_menu'
    return 'school_number'

def main_menu(update: Update, context: CallbackContext):

    data = update.message.text
    if data.isnumeric():
        context.user_data['class_num'] = update.message.text
        cursor.execute(query, (context.user_data['name'],
                               context.user_data['surname'],
                               context.user_data['father_name'],
                               context.user_data['id'],
                               context.user_data['number'],
                               context.user_data['region'],
                               context.user_data['district'],
                               context.user_data['school_num'],
                               context.user_data['class_num']
                               ))
        connection.commit()
        # connection.close()
        update.message.reply_html("‹‹ ʙɪʟɪᴍᴛᴇsᴛʙᴏᴛ ››", reply_markup=ReplyKeyboardMarkup([["🔠 Test Ishlash"], [" ℹ️ Ma'lumot", "👤 Profilim"]], resize_keyboard=True))
    
        # context.user_data['number'] = update.message.contact.phone_number
    # work_sheet.append([f"{context.user_data['familiya']}", f"{context.user_data['ism']}", f"{context.user_data['sharif']}", f"{context.user_data['number']}", f"{context.user_data['id']}"])
    # cursor.execute(query, (context.user_data['ism'] , context.user_data['familiya'], context.user_data['sharif'], context.user_data['id'], context.user_data['number'], 'Samarqand', 'Oqdaryo', 12, 11))

    # wb.save('users.xlsx')
   

def get_user_list(update: Update, context: CallbackContext):
    
    update.effective_chat.send_action('upload_document')

    # if update.effective_user.id == owner_id:
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    cursor.execute("""SELECT * FROM users WHERE user_id = user_id""")
    # update.message.reply_text()            work_sheet.append([ f"{cur[1]}", f"{cur[2]}", f"{cur[3]}", f"{cur[4]}", f"{cur[5]}", f"{cur[6]}", f"{cur[7]}", f"{cur[8]}", f"{cur[9]}"])
    work_sheet.append([ f"Ism", f"Familiya", f"Sharif", f"User_id", f"Tel_raqam", f"Viloyat", f"Tuman", f"Maktab_raqami", f"Sinf_raqami"])

    for cur in cursor.fetchall():
            work_sheet.append([ f"{cur[1]}", f"{cur[2]}", f"{cur[3]}", f"{cur[4]}", f"{cur[5]}", f"{cur[6]}", f"{cur[7]}", f"{cur[8]}", f"{cur[9]}"])
            wb.save('users.xlsx')    
    doc_file = open('users.xlsx', 'rb')
 
    return context.bot.send_document(chat_id, doc_file)


work_sheet.delete_rows(1, work_sheet.max_row+1) # for delete entire sheet
wb.save('users.xlsx')

def get_user_info(update: Update, context: CallbackContext):
    user_id  = update.effective_user.id
    cursor.execute(f'''SELECT * FROM users  WHERE user_id = {user_id} ''')
    
    for cur in cursor:
        info_list = cur
    text = "<b>Sizning ma'lumotlaringiz:\n\n</b>"
    text += f"<b>F.I.SH.</b>: {info_list[3]} {info_list[1]} {info_list[2]}\n"
    text += f"<b>Telefon raqamingiz</b>: {info_list[5]}\n"
    text += f"<b>Viloyat</b>: {info_list[6]}\n"
    text += f"<b>Tuman</b>: {info_list[7]}\n"
    text += f"<b>Maktab</b>: {info_list[8]}\n"
    text += f"<b>Sinf</b>: {info_list[9]}\n"
    update.message.reply_html(text)
    
 
updater.dispatcher.add_handler(MessageHandler(filters=Filters.text('give_user_list'), callback=get_user_list))
updater.dispatcher.add_handler(MessageHandler(filters=Filters.text("👤 Profilim"), callback=get_user_info)) 
updater.dispatcher.add_handler(ConversationHandler(
                                entry_points=[CommandHandler('start', start_bot)],
                                            #   CommandHandler('start', start)],
                                states={
                                    'fullname':[MessageHandler(filters=Filters.update, callback=get_fullname1)],
                                    'repeat_name':[MessageHandler(filters=Filters.update, callback=get_fullname2)],
                                    'phone_number':[MessageHandler(filters=Filters.update, callback = get_contact),
                                                    MessageHandler(filters=Filters.contact, callback = get_contact)],
                                    'region':[MessageHandler(filters=Filters.update, callback=get_region)],
                                    'district':[MessageHandler(filters=Filters.text(districts), callback=get_district)],
                                    'school_number':[MessageHandler(filters=Filters.update, callback=get_school_number)],
                                    'class_number':[MessageHandler(filters=Filters.update, callback=get_class_number)],
                                    'main_menu':[MessageHandler(filters=Filters.update, callback=main_menu)],
                                    
                                    
                                    
                                                        
                                },
                                fallbacks=[ 
                                            MessageHandler(filters=Filters.command, callback=start_bot),
                                            MessageHandler(filters=Filters.text('⬅️ Orqaga'),callback=start_bot)]
))  

 
    

    
# updater.dispatcher.add_handler(CommandHandler('start', start_bot))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text("📝 Ro'yxatdan o'tish"), callback=get_fullname1))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.contact , callback=get_region))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text("⬅️ Orqaga"), callback=start_bot))
# updater.dispatcher.add_handler(MessageHandler( filters=filter_fullname, callback=get_contact))
# updater.dispatcher.add_handler(MessageHandler( filters=filter_number, callback=get_region))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text(regions), callback=get_district))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text("👤 Profilim"), callback=get_user_list))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text(districts), callback=get_school_number))
# updater.dispatcher.add_handler(MessageHandler(filters=Filters.filter_number))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text, callback=get_fullname2))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text("Location"), callback=get_location))
updater.start_polling()
updater.idle()
connection.close()



  # region_list = [list([cur[0]]) for cur in cursor]
    # K = region_list.append(['⬅️ Orqaga'])
    # return get_school_number






# def is_user_there(update, context, channel, chat_id):
    #     try:
#        chat_member = context.bot.get_chat_member(channel,chat_id) # this will return a ChatMember object
#        if chat_member.status in ['administrator', 'creator', 'member']:
#          return True
#        else:
#          return False
#     except BadRequest:
#        return False





  # chat_member = context.bot.get_chat_member(chat_id = '@tuitgr_20', user_id = update.effective_user.id) # this will return a ChatMember object
    #    logging.error(chat_member)
    # print("\n\n\n\n\n\n\n\n\n\n\n\n\n",chat_member, "\n\n\n\n\n\n\n\n\n\n\n\n\n")
    # update.effective_chat.send_message()
    # try:
 
    #    chat_member = context.bot.get_chat_member(chat_id = "@kunuz", user_id = update.effective_user.id) # this will return a ChatMember object
    # #    logging.error(chat_member)
    #    print("\n\n\n\n\n\n\n\n\n\n\n\n\n",chat_member, "\n\n\n\n\n\n\n\n\n\n\n\n\n")
    #    if chat_member.status in ['administrator', 'member', "creator"]:
    #     #  update.message.reply_text(chat_member.status, chat_member.user)
    #      update.message.reply_text("ISHLADI")
    #      return True

    #    else:
    #      update.message.reply_text("ISHLAMADI1")
    #      return False
    # except Exception:
    #    update.message.reply_text("ISHLAMADI2")
    #    return False

    # context.bot.get_chat_member(chat_id="@kunuz", user_id=int(user_id))
    # update.message.reply_text(chat_member)


     
    # first_column = work_sheet['E']
    # user_list = []
    # for x in range(len(first_column)): 
    #     print(first_column[x].value, type(first_column[x].value))
    #     if str(first_column[x].value).isnumeric():
    #         user_list.append(int(first_column[x].value))
    # logging.debug(user_list,"\n\n\n\n\n\n")
    # if update.effective_user.id in user_list:
    #     update.message.reply_text("Quyidagi menyulardan birini tanlang", reply_markup=ReplyKeyboardMarkup([["🔠 Test Ishlash"], [" ℹ️ Ma'lumot", "👤 Profilim"]], resize_keyboard=True))
    # else:










    # bot = Bot(token="5200114262:AAFO9hhfgCQL1FjQYOxBF0l2iILz5pz16EE")
    # chat_member = bot.get_chat_member(chat_id="@kunuz", user_id=int(user_id))
    # logging.debug(chat_member)
    # if is_user_there(update, context, "@kunuz", chat_id=int(user_id)):
    #     update.message.reply_text("ISHLADI")
    # else:
        # update.message.reply_text("ISHLAMADI")
    # bot  = Bot(token="5200114262:AAFO9hhfgCQL1FjQYOxBF0l2iILz5pz16EE")
    # result = bot.get_chat_member(chat_id=-1001167945861, user_id = int(update.effective_user.id))
    # update.message.reply_text(result.status)
    






# def get_location(update: Update, context: CallbackContext):
 
#     key = ReplyKeyboardMarkup([[KeyboardButton("📍 Joylashgan manzilni jo'natish", request_location=True)]],
#                                  resize_keyboard=True)
#     update.message.reply_text("Manzilni yuborish uchun quyidagi tugmani bosing", reply_markup=key)


 

# if update.message.text == '📝 Ro\'yxatdan o\'tish':



# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text(), callback=check_fullname))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text(["F.I.SH. kiriting", "Familiya, ism, sharifingizni to'g'ri kiriting"]), callback=get_fullname))
# updater.dispatcher.add_handler(MessageHandler( filters=Filters.text("Familiya, ism, sharifingizni to'g'ri kiriting"), callback=check_fullname))

# BUTTON2 = ReplyKeyboardMarkup([["📱 Telefon raqamni jo'natish"]], resize_keyboard=True)
 
# updater.dispatcher.add_handler(ConversationHandler(
#                                 entry_points=[CommandHandler('start', start_bot)],
#                                 states={
#                                     '1':[MessageHandler(filters=Filters.text("Ro'yxatdan o'tish"), callback=send_data)]
#                                 },
#                                 fallbacks=[MessageHandler(filters=Filters.command, callback=cancel)]
# ))
# def cancel(update: Update, context: CallbackContext):
#     """ Cancel current conversation """
#     update.message.reply_text('Siz noto\'g\'ri buyruqni belgiladingiz\n'\
#                               'Botni ishga tushirish uchun /start ni bosing')
#     return ConversationHandler.END 



# def check_fullname(update: Update, context: CallbackContext):
#     # text = ism familiya 
    
#     text = update.message.text
#     update.message.reply_text(text+" "+f"{text.split()}")
#     if len(text.split()) > 1:
#         if text.split()[0][0].isupper() and text.split()[1][0].isupper():
#             update.message.reply_text('Tugmani bosib telefon raqamingizni yuboring')
#         else:
#             update.message.reply_text("Familiya, ism, sharifingizni to'g'ri kiriting")
    
            